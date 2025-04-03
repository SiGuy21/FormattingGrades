'''
team 9
section 03
This formats grades in excel
'''
import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font

inputFile = "Poorly_Organized_Data_1.xlsx"
inputWorkBook = openpyxl.load_workbook(inputFile)
inputSheet = inputWorkBook.active

myWorkbook = Workbook()
myWorkbook.remove(myWorkbook["Sheet"])


# 1. Sheets for each class

headers = ["Last Name", "First Name", "Student ID", "Grade"]
subjects = []
subject_sheets = {}

# Create a new sheet for each unique class and add headers
for row in inputSheet.iter_rows(min_row=2, max_col=1, values_only=True):
    className = row[0]
    if className not in subjects:
        subjects.append(className)
        subject_sheets[className] = myWorkbook.create_sheet(className)
        subject_sheets[className].append(headers)
        print(f"Added {className} sheet to workbook")

# 2. last name, first name, student ID, and grade columns


# Iterate through the input sheet and add data to the corresponding class sheet
for row in inputSheet.iter_rows(min_row = 2, values_only=True):
    className = row[0]
    studentInfo = row[1]
    grade = row[2]
    # Split studentInfo into last name, first name, and student ID
    lastName, firstName, studentID = studentInfo.split('_')
    studentData = [lastName, firstName, studentID, grade]
    # Append the student data to the corresponding class sheet
    if className in subject_sheets:
        subject_sheets[className].append(studentData)
        print(f"Added data for {firstName} {lastName} to {className} sheet")


# 3. Filter
for className, sheet in subject_sheets.items():
    # Get the maximum row number with data
    max_row = sheet.max_row
    # Apply filter to range A1:D{max_row}
    sheet.auto_filter.ref = f"A1:D{max_row}"
    print(f"Applied filter to {className} sheet")

# 4. Adding functions 
for className, sheet in subject_sheets.items():
    sheet["F1"] = "Summary Statistics"
    sheet["F2"] = "Highest Grade"
    sheet["F3"] = "Lowest Grade"
    sheet["F4"] = "Mean Grade"
    sheet["F5"] = "Median Grade"
    sheet["F6"] = "Number of Students"
    sheet["G1"] = "Value"
    sheet["G2"] = "=MAX(D2:D41)"
    sheet["G3"] = "=MIN(D2:D41)"
    sheet["G4"] = "=AVERAGE(D2:D41)"
    sheet["G5"] = "=MEDIAN(D2:D41)"
    sheet["G6"] = "=COUNT(D2:D41)"

# 5. Simple formatting
for className, sheet in subject_sheets.items():
    # Boldings
    # Bold the first row (headers) of Columns A thru D and F thru G
    # Skip E since it's blank column for spacing
    for col in range(1, 5):
        sheet.cell(row = 1, column = col).font = Font(bold = True)
    for col in range(6, 8):
        sheet.cell(row = 1, column = col).font = Font(bold = True)
    
    # Adjusting widths
    '''
    # Hard-coded version
    column_widths = {"A": 14, "B": 14, "C": 13, "D": 10, "F": 15, "G": 15}
    for iCol, width in column_widths.items():
        sheet.column_dimensions[iCol].width = width
    ''' 
    # Adjusts column widths based on header length (+5)
    for colI, header in enumerate(headers, start = 1):
        colL = openpyxl.utils.get_column_letter(colI)
        sheet.column_dimensions[colL].width = len(header) + 5
        
    # Make sure Columns F and G also gets their widths adjusted as well
    sheet.column_dimensions["F"].width = 23
    sheet.column_dimensions["G"].width = 10

# 6. Save the results
myWorkbook.save(filename = "formatted_grades.xlsx")
myWorkbook.close()
